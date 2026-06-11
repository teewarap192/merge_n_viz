import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference

HIGHLIGHT_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
VERTICAL_ALIGN = Alignment(textRotation=90, horizontal='center', vertical='bottom')

STEP_MAPPING = {
    'DS4114': 'DS4114 / S1',
    'DS4214': 'DS4214 / S2',
    'DS4314': 'DS4314 / S3'
}

def ensure_extension(filename):
    """Automatically adds .xlsx if you forget to type it."""
    if not filename.lower().endswith('.xlsx'):
        return filename + '.xlsx'
    return filename

def update_master_file():
    print("--- Last Wafer Monitoring Update Tool ---")
    
    # --- Dynamic Inputs ---
    week_num = input("1. Enter the current week number (e.g., 35): ").strip()
    if not week_num:
        print("Operation cancelled.")
        return
    new_prefix = f"LW{week_num}-"

    file1_input = input("2. Enter the name of the NEW data file (e.g., file1): ").strip()
    file2_input = input("3. Enter the name of the MASTER data file (e.g., file2): ").strip()
    output_input = input("4. Enter the name to SAVE AS (e.g., final_report): ").strip()

    file1_path = ensure_extension(file1_input)
    file2_path = ensure_extension(file2_input)
    output_path = ensure_extension(output_input)

    print(f"\nLoading master database from {file2_path}...")
    try:
        wb = load_workbook(file2_path)
    except FileNotFoundError:
        print(f"\nError: Could not find '{file2_path}'. Please check the name and ensure it is in the same folder.")
        return
        
    ws = wb.active

    # --- PHASE 1: Merge Previous Week's Data & Clean Up ---
    print("Housekeeping: Merging old week data and cleaning columns...")

    base_cols = {}
    max_col_initial = ws.max_column
    for c in range(3, max_col_initial + 1):
        val = ws.cell(row=2, column=c).value
        if val and not re.match(r'^LW\d+-', str(val).strip()):
            base_cols[str(val).strip()] = c

    cols_to_delete = []

    for c in range(max_col_initial, 2, -1):
        header_val = ws.cell(row=2, column=c).value
        if header_val:
            header_str = str(header_val).strip()

            if re.match(r'^LW\d+-', header_str):
                base_lot = re.sub(r'^LW\d+-', '', header_str)

                if base_lot in base_cols:
                    base_c = base_cols[base_lot]
                    for r in range(4, ws.max_row + 1):
                        cell_val = ws.cell(row=r, column=c).value
                        if cell_val is not None and str(cell_val).strip() != "":
                            ws.cell(row=r, column=base_c).value = cell_val
                    cols_to_delete.append(c)
                else:
                    ws.cell(row=2, column=c).value = base_lot
                    base_cols[base_lot] = c

    for c in cols_to_delete:
        ws.delete_cols(c)

    print("Resetting all background colors...")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = NO_FILL

    # --- PHASE 2: Import Current Week's Data ---
    print(f"Reading new data from {file1_path}...")

    try:
        df_raw = pd.read_excel(file1_path, header=None)
    except FileNotFoundError:
        print(f"\nError: Could not find '{file1_path}'. Please check the name and ensure it is in the same folder.")
        return

    header_idx = -1
    for idx, row in df_raw.iterrows():
        if 'Lot' in row.values:
            header_idx = idx
            break

    if header_idx == -1:
        print(f"Error: Could not find the 'Lot' header in {file1_path}.")
        return

    df_new = pd.read_excel(file1_path, header=header_idx)

    step_rows = {}
    for r in range(1, ws.max_row + 5):
        val = ws.cell(row=r, column=2).value
        if val:
            step_rows[str(val).strip()] = r

    if 'Target' not in step_rows:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=2, value="Target")
        step_rows['Target'] = target_row

    cells_to_highlight = []
    current_max_col = ws.max_column

    print(f"Staging new data for Week {week_num}...")

    for index, row_data in df_new.iterrows():
        if 'Lot' not in row_data or pd.isna(row_data['Lot']):
            continue

        lot_val = str(row_data['Lot']).strip()
        if not lot_val or lot_val == 'nan':
            continue

        new_lot_name = f"{new_prefix}{lot_val}"

        current_max_col += 1
        target_col = current_max_col
        ws.cell(row=2, column=target_col, value=new_lot_name)
        cells_to_highlight.append(ws.cell(row=2, column=target_col))

        ws.cell(row=step_rows['Target'], column=target_col, value=1)

        for file1_step, file2_step in STEP_MAPPING.items():
            if file1_step in row_data:
                val_to_copy = row_data[file1_step]

                if pd.notna(val_to_copy) and str(val_to_copy).strip() != "":
                    if file2_step in step_rows:
                        target_row = step_rows[file2_step]
                    else:
                        target_row = ws.max_row + 1
                        ws.cell(row=target_row, column=2, value=file2_step)
                        step_rows[file2_step] = target_row

                    cell = ws.cell(row=target_row, column=target_col)
                    cell.value = val_to_copy
                    cells_to_highlight.append(cell)

    # --- PHASE 3: Formatting & Chart Generation ---
    print("Applying highlights and vertical formatting...")
    for cell in cells_to_highlight:
        cell.fill = HIGHLIGHT_FILL

    for c in range(3, current_max_col + 1):
        ws.cell(row=2, column=c).alignment = VERTICAL_ALIGN

    print("Generating Chart for the last 80 lots...")
    ws._charts.clear()

    min_c = max(3, current_max_col - 79) 

    r_s1 = step_rows.get('DS4114 / S1')
    r_s2 = step_rows.get('DS4214 / S2')
    r_s3 = step_rows.get('DS4314 / S3')
    r_target = step_rows.get('Target')

    if r_s1 and r_s2 and r_s3 and r_target:
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "Overall Defect Monitoring from Last Wafer Monitoring process"
        
        bar_chart.height = 20
        bar_chart.width = 45 
        bar_chart.legend.position = "b" 
        
        bar_chart.y_axis.title = "Overall defect rate (%)"
        bar_chart.y_axis.tickLblPos = "nextTo"
        bar_chart.y_axis.number_format = '0.00' 
        bar_chart.y_axis.majorUnit = 0.2
        bar_chart.y_axis.scaling.max = 1.4 
        bar_chart.y_axis.scaling.min = 0.0

        bar_chart.x_axis.tickLblPos = "low"
        bar_chart.x_axis.tickLblSkip = 1

        for r_idx in [r_s1, r_s2, r_s3]:
            data = Reference(ws, min_col=min_c, max_col=current_max_col, min_row=r_idx, max_row=r_idx)
            s = openpyxl.chart.Series(data, title=ws.cell(row=r_idx, column=2).value)
            bar_chart.series.append(s)

        cats = Reference(ws, min_col=min_c, max_col=current_max_col, min_row=2, max_row=2)
        bar_chart.set_categories(cats)

        line_chart = LineChart()
        t_data = Reference(ws, min_col=min_c, max_col=current_max_col, min_row=r_target, max_row=r_target)
        t_series = openpyxl.chart.Series(t_data, title="Target")
        
        t_series.graphicalProperties.line.solidFill = "FF0000" 
        t_series.graphicalProperties.line.width = 25000 
        line_chart.series.append(t_series)

        bar_chart += line_chart

        ws.add_chart(bar_chart, "B13")

    wb.save(output_path)
    print(f"\nProcess complete! Formats cleaned, Chart Updated, and saved successfully as '{output_path}'.")

if __name__ == "__main__":
    update_master_file()